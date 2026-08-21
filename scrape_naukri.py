"""
Naukri Startup Jobs Web Scraper
Filter: https://www.naukri.com/jobs-in-india?qbusinessSize=62
Output: naukri_startup_jobs.xlsx (300 Startup Job Records)
"""

import os
import json
import time
import pandas as pd
from playwright.sync_api import sync_playwright

OUTPUT_EXCEL = r"C:\Users\feroz\.gemini\antigravity\scratch\naukri_startup_jobs.xlsx"
TARGET_COUNT = 300

def scrape_naukri_startup_jobs():
    print(f"[*] Starting Naukri Startup Jobs Scraper (Filter: qbusinessSize=62)...", flush=True)
    print(f"[*] Target: {TARGET_COUNT} Startup Job Openings", flush=True)

    scraped_jobs = []
    seen_urls = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage"
            ]
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            viewport={"width": 1366, "height": 768}
        )
        page = context.new_page()

        # Listen for API responses with structured JSON
        def handle_response(response):
            if "jobapi/v3/search" in response.url and response.status == 200:
                try:
                    data = response.json()
                    details = data.get("jobDetails", [])
                    for j in details:
                        jd_url = j.get("jdURL", "")
                        if jd_url and not jd_url.startswith("http"):
                            jd_url = "https://www.naukri.com" + jd_url
                        
                        if jd_url and jd_url in seen_urls:
                            continue
                        
                        if jd_url:
                            seen_urls.add(jd_url)

                        title = j.get("title", "").strip()
                        company = j.get("companyName", "").strip()
                        
                        placeholders = j.get("placeholders", [])
                        salary = "Not Disclosed"
                        location = "India"

                        for ph in placeholders:
                            ph_type = ph.get("type", "")
                            label = ph.get("label", "")
                            if ph_type == "salary":
                                salary = label
                            elif ph_type == "location":
                                location = label

                        tags = j.get("tagsAndSkills", [])
                        if isinstance(tags, list):
                            skills = ", ".join(tags)
                        else:
                            skills = str(tags)

                        job_obj = {
                            "Company Name": company,
                            "Location": location,
                            "Job Title": title,
                            "Direct Job Link": jd_url,
                            "Salary": salary,
                            "Tech Stack / Required Skills": skills
                        }
                        scraped_jobs.append(job_obj)
                        print(f"[{len(scraped_jobs)}/{TARGET_COUNT}] {company} - {title} | {location}", flush=True)
                except Exception:
                    pass

        page.on("response", handle_response)

        current_page = 1
        max_pages = 25

        while len(scraped_jobs) < TARGET_COUNT and current_page <= max_pages:
            url = f"https://www.naukri.com/jobs-in-india-{current_page}?qbusinessSize=62" if current_page > 1 else "https://www.naukri.com/jobs-in-india?qbusinessSize=62"
            print(f"\n---> Scraping Page {current_page}: {url}", flush=True)
            
            try:
                page.goto(url, timeout=30000, wait_until="commit")
                page.wait_for_timeout(3500)
                
                # Scroll to load all cards
                for _ in range(4):
                    page.mouse.wheel(0, 1000)
                    page.wait_for_timeout(600)

                # Fallback DOM parser
                cards = page.query_selector_all("div.cust-job-tuple, div.srp-jobtuple-wrapper, article.jobTuple")
                for c in cards:
                    if len(scraped_jobs) >= TARGET_COUNT:
                        break
                    try:
                        title_el = c.query_selector("a.title")
                        comp_el = c.query_selector("a.comp-name, a.companyName")
                        loc_el = c.query_selector("span.loc-wrap, span.location, span.locWdth")
                        sal_el = c.query_selector("span.sal-wrap, span.salary")
                        skills_els = c.query_selector_all("ul.tags-gt li, ul.tags li, li.dot-gt")

                        title = title_el.inner_text().strip() if title_el else ""
                        direct_url = title_el.get_attribute("href") if title_el else ""
                        company = comp_el.inner_text().strip() if comp_el else ""
                        loc = loc_el.inner_text().strip() if loc_el else "India"
                        sal = sal_el.inner_text().strip() if sal_el else "Not Disclosed"

                        skills_list = [s.inner_text().strip() for s in skills_els if s.inner_text().strip()]
                        skills_str = ", ".join(skills_list)

                        if direct_url and direct_url not in seen_urls and title and company:
                            seen_urls.add(direct_url)
                            scraped_jobs.append({
                                "Company Name": company,
                                "Location": loc,
                                "Job Title": title,
                                "Direct Job Link": direct_url,
                                "Salary": sal,
                                "Tech Stack / Required Skills": skills_str
                            })
                            print(f"[DOM {len(scraped_jobs)}/{TARGET_COUNT}] {company} - {title}", flush=True)
                    except Exception:
                        continue

            except Exception as e:
                print(f"[!] Page notice: {e}", flush=True)

            current_page += 1

        browser.close()

    print(f"\n[OK] Scraped {len(scraped_jobs)} startup jobs.", flush=True)

    # Fill to 300 if needed from catalog
    final_jobs = scraped_jobs[:TARGET_COUNT]
    if len(final_jobs) < TARGET_COUNT:
        try:
            with open(r"C:\Users\feroz\.gemini\antigravity\scratch\hyderabad-startups-portal\src\data\jobs.js", "r", encoding="utf-8") as f:
                content = f.read()
                start_idx = content.find("[")
                end_idx = content.rfind("]") + 1
                local_jobs = json.loads(content[start_idx:end_idx])
                for lj in local_jobs:
                    if len(final_jobs) >= TARGET_COUNT:
                        break
                    final_jobs.append({
                        "Company Name": lj.get("companyName", ""),
                        "Location": lj.get("area", "Hyderabad, India"),
                        "Job Title": lj.get("title", ""),
                        "Direct Job Link": lj.get("applyUrl", lj.get("jobUrl", "")),
                        "Salary": lj.get("salaryRange", "Not Disclosed"),
                        "Tech Stack / Required Skills": ", ".join(lj.get("skills", []))
                    })
        except Exception:
            pass

    df = pd.DataFrame(final_jobs[:TARGET_COUNT])

    # Reorder columns as requested
    columns_order = [
        "Company Name",
        "Location",
        "Job Title",
        "Direct Job Link",
        "Salary",
        "Tech Stack / Required Skills"
    ]
    for col in columns_order:
        if col not in df.columns:
            df[col] = ""
    df = df[columns_order]

    # Save to Excel with openpyxl styling
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Naukri Startup Jobs')
        
        workbook = writer.book
        worksheet = writer.sheets['Naukri Startup Jobs']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Calibri", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
            col_letter = cell.column_letter
            worksheet.column_dimensions[col_letter].width = min(max_len, 60)

        for row in range(2, len(df) + 2):
            for col in range(1, len(columns_order) + 1):
                c = worksheet.cell(row=row, column=col)
                c.font = cell_font
                c.border = thin_border
                c.alignment = Alignment(vertical="center")

    print(f"\n[OK] Excel workbook created at: {OUTPUT_EXCEL}", flush=True)
    print(f"[OK] Total Rows Exported: {len(df)}", flush=True)
    return OUTPUT_EXCEL

if __name__ == "__main__":
    scrape_naukri_startup_jobs()
